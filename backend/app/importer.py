"""Excel/CSV import: Persian header mapping, upsert, account creation."""
import csv
import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .core.persian import light_normalize
from .models import ChangeLog, Employee

# Persian (and common alternate) headers -> model field
HEADER_MAP = {
    "نام": "first_name",
    "نام همکار": "first_name",
    "همکار": "first_name",
    "firstname": "first_name",
    "نام خانوادگی": "last_name",
    "lastname": "last_name",
    "نام لاتین": "latin_name",
    "لاتین": "latin_name",
    # خروجی خام سانترال — فقط همین سه ستون؛ بقیه (از جمله Secret) رمز و تنظیمات‌اند.
    "display name": "latin_name",
    "user extension": "extension",
    "direct did": "direct_number",
    "شماره مستقیم": "direct_number",
    "خط مستقیم": "direct_number",
    "مستقیم": "direct_number",
    "شماره داخلی (اداری)": "extension",
    "داخلی": "extension",
    "شماره داخلی": "extension",
    "شماره ثابت": "phone",
    "تلفن ثابت": "phone",
    "ایمیل سازمانی": "email",
    "ایمیل": "email",
    "واحد / دپارتمان": "department",
    "واحد": "department",
    "دپارتمان": "department",
    "شرکت": "company",
    "سمت": "job_title",
    "محل خدمت": "location",
    "آدرس": "location",
    "عکس (مسیر/لینک)": "photo_url",
    "عکس": "photo_url",
    "کلمات کلیدی و نام\u200cهای مستعار": "keywords",
    "کلیدواژه": "keywords",
    "کلمات کلیدی": "keywords",
    "مهارت\u200cها / حوزه کاری": "skills",
    "مهارت‌ها": "skills",
    "زبان\u200cها": "languages",
    "زبان": "languages",
    "ساعت کاری / بهترین زمان تماس": "working_hours",
    "ساعت کاری": "working_hours",
    "یادداشت": "notes",
}

# columns that identify a row for upsert purposes
KEY_FIELDS = ("email", "extension", "direct_number", "latin_name")

# ستون‌هایی که بدون آن‌ها سطر پذیرفته نمی‌شود
REQUIRED_FIELDS = {
    "first_name": "نام",
    "last_name": "نام خانوادگی",
    "department": "واحد",
    "job_title": "سمت",
    "extension": "داخلی",
}


def _cell_str(v) -> str:
    if v is None:
        return ""
    return light_normalize(str(v)).strip()


def _norm_header(h) -> str:
    if h is None:
        return ""
    return light_normalize(str(h)).replace("\u200c", "").replace(" ", "").lower()


_NORMED_MAP = None


def _normed_map() -> dict[str, str]:
    global _NORMED_MAP
    if _NORMED_MAP is None:
        _NORMED_MAP = {(_norm_header(raw) or "\x00"): field for raw, field in HEADER_MAP.items()}
    return _NORMED_MAP


def _match_field(nh: str) -> str | None:
    """Exact match first; then longest-header containment (guarded).

    Containment requires the mapped header to be >=4 normalized chars so
    short headers like 'نام' can't hijack longer ones like
    'کلمات کلیدی و نام‌های مستعار'.
    """
    m = _normed_map()
    if nh in m:
        return m[nh]
    best_raw, best_field = None, None
    for nraw, field in m.items():
        if nraw == "\x00":
            continue
        if len(nraw) >= 4 and nraw in nh:
            if best_raw is None or len(nraw) > len(best_raw):
                best_raw, best_field = nraw, field
    return best_field


def _blocks(headers) -> list[dict[int, str]]:
    """ستون‌ها را به جدول‌های کنارهم می‌شکند.

    تکرارِ یک فیلد در سرستون یعنی جدولِ بعدی شروع شده — «واحد|داخلی|واحد|داخلی»
    دو جدول است، نه چهار ستونِ یک جدول.
    """
    out, cur = [], {}
    for i, h in enumerate(headers):
        field = _match_field(_norm_header(h))
        if not field:
            continue
        if field in cur.values():
            out.append(cur)
            cur = {}
        cur[i] = field
    if cur:
        out.append(cur)
    return out


def _split_extensions(ext: str) -> list[str]:
    """«۱۰۲-۱۰۳» دو داخلی است، نه یک شماره‌ی بلند."""
    parts = [p.strip() for p in ext.split("-")]
    if len(parts) > 1 and all(p.isdigit() and len(p) <= 5 for p in parts):
        return parts
    return [ext]


def parse_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return _parse_rows(rows)


def parse_csv(content: bytes) -> list[dict]:
    """جداکننده حدس زده می‌شود — همه‌ی خروجی‌های تلفن‌سانترال کاما نیستند."""
    text = content.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|\x1b")
    except csv.Error:
        dialect = csv.excel
    return _parse_rows(list(csv.reader(io.StringIO(text), dialect)))


def _parse_rows(rows) -> list[dict]:
    if not rows:
        return []
    # سرستون همیشه سطر اول نیست: بالایش سطر خالی یا عنوان می‌گذارند.
    # سطری که بیشترین ستونِ شناخته‌شده را بدهد سرستون است.
    hdr, blocks = 0, []
    for i, row in enumerate(rows[:20]):
        b = _blocks(row)
        if sum(map(len, b)) > sum(map(len, blocks)):
            hdr, blocks = i, b
    records = []
    section = [""] * len(blocks)
    for row in rows[hdr + 1:]:
        for bi, colmap in enumerate(blocks):
            rec = {}
            for idx, field in colmap.items():
                if idx < len(row):
                    rec[field] = _cell_str(row[idx])
            if not any(v for v in rec.values()):
                continue
            # سطری که فقط واحد دارد عنوانِ بخش است، نه یک نفر:
            # واحدِ سطرهای زیرِ خودش می‌شود و برچسبِ خودش سمت می‌شود.
            if rec.get("department") and not any(v for k, v in rec.items() if k != "department"):
                section[bi] = rec["department"]
                continue
            if section[bi] and rec.get("department"):
                if not rec.get("job_title"):
                    rec["job_title"] = rec["department"]
                rec["department"] = section[bi]
            parts = _split_extensions(rec.get("extension", ""))
            if len(parts) > 1:
                records.extend({**rec, "extension": p} for p in parts)
            else:
                records.append(rec)
    return records


def _find_existing(db: Session, rec: dict, organization_id: int) -> Employee | None:
    for f in KEY_FIELDS:
        val = (rec.get(f) or "").strip()
        if not val:
            continue
        col = getattr(Employee, f)
        emp = db.query(Employee).filter(
            Employee.organization_id == organization_id, col == val
        ).first()
        if emp:
            return emp
    fn, ln = rec.get("first_name", ""), rec.get("last_name", "")
    if fn and ln:
        emp = (
            db.query(Employee)
            .filter(Employee.organization_id == organization_id,
                    Employee.first_name == fn, Employee.last_name == ln)
            .first()
        )
        if emp:
            return emp
    return None


def apply_import(db: Session, records: list[dict], actor, organization_id: int | None = None) -> dict:
    """ایمپورت فقط دفترچه را پر می‌کند — حسابِ ورود نمی‌سازد.

    دفترچه ورود نمی‌خواهد، پس پرسنل حساب لازم ندارند؛ حساب فقط برای
    ادمین است و دستی ساخته می‌شود.
    """
    organization_id = organization_id or actor.organization_id
    created_n = updated_n = skipped_n = 0
    errors: list[str] = []

    for i, rec in enumerate(records, start=2):  # excel row numbers (header=row1)
        try:
            name = " ".join(x for x in [rec.get("first_name", ""), rec.get("last_name", "")] if x)
            missing = [label for f, label in REQUIRED_FIELDS.items() if not (rec.get(f) or "").strip()]
            if missing:
                skipped_n += 1
                errors.append(f"سطر {i}: ستون‌های اجباری خالی است: {'، '.join(missing)}")
                continue

            emp = _find_existing(db, rec, organization_id)
            if emp is None:
                emp = Employee(organization_id=organization_id)
                db.add(emp)
                created_n += 1
                action = "create"
            else:
                updated_n += 1
                action = "update"

            for field in set(HEADER_MAP.values()):
                val = rec.get(field)
                if val is not None and val != "":
                    setattr(emp, field, val)
            emp.sync_direct_number()
            emp.rebuild_search_text()
            # flush now so a freshly-created employee gets its id
            # (the change log below needs it)
            db.flush()
            db.add(
                ChangeLog(
                    organization_id=organization_id,
                    entity="employee",
                    entity_id=emp.id,
                    action="import" if action == "create" else "import_update",
                    actor_id=getattr(actor, "id", None),
                    actor_name=getattr(actor, "username", "system"),
                    actor_role=getattr(actor, "role", None),
                    details={"row": i, "name": light_normalize(name)},
                )
            )
        except Exception as exc:  # keep importing other rows
            skipped_n += 1
            errors.append(f"سطر {i}: {exc}")

    db.commit()
    try:
        from . import fts as _fts
        _fts.reindex_all(db)
    except Exception:
        pass  # FTS اختیاری است؛ init_db در restart هم rebuild می‌کند
    return {
        "created": created_n,
        "updated": updated_n,
        "skipped": skipped_n,
        "errors": errors,
    }
